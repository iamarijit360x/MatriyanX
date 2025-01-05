from openpyxl.styles import Border, Side,PatternFill,Font,Alignment
from openpyxl import Workbook
from io import BytesIO
class ExcelService:
    def __init__(self):
        self.name = "Excel Class"
    def data_sanitization(self, sample_data):
        newdata=[]
        i=0
        while(i<len(sample_data)):
            temp={}
            if(i!=len(sample_data)-1 and sample_data[i]['name']==sample_data[i+1]['name'] and sample_data[i]['date']==sample_data[i+1]['date']):
                
                temp['name']=sample_data[i]['name']
                temp['village']=sample_data[i]['village']
                token=[]
                distance=[]
                date=[]
                token.append(sample_data[i]['voucher_type'])
                token.append(sample_data[i+1]['voucher_type'])
                distance.append(sample_data[i]['distance'])
                distance.append(sample_data[i+1]['distance'])
                date.append(sample_data[i]['date'])
                date.append(sample_data[i+1]['date'])
                temp['voucher_type']=token
                temp['distance']=distance
                temp['date']=date
                newdata.append(temp)
                i+=1
            else:
                newdata.append(sample_data[i])
            i+=1
        return newdata
    
    def add_borders_top(self,sheet, start_row=1, end_row=9, columns=['A','B','C','D','E','F','G','H']):
        """
        Add black borders to specified columns and rows in the given worksheet.

        Parameters:
        - sheet: openpyxl.worksheet.worksheet.Worksheet, the worksheet where borders will be added.
        - start_row: int, the starting row number for adding borders.
        - end_row: int, the ending row number for adding borders.
        - columns: list of str, column letters where borders will be applied.
        """
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        for row in range(start_row, end_row + 1):
            for col in columns:
                cell = sheet[f'{col}{row}']
                cell.border = border

    # Example usage
    def add_borders(self,sheet, start_row, end_row, columns):
        """
        Add custom borders to specified columns and rows in the given worksheet.
        
        Parameters:
        - sheet: openpyxl.worksheet.worksheet.Worksheet, the worksheet where borders will be added.
        - start_row: int, the starting row number for adding borders.
        - end_row: int, the ending row number for adding borders.
        - columns: list of str, column letters where borders will be applied.
        """
        # Define border styles
        top_border = Border(top=Side(border_style="thin", color="000000"),left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"))
        bottom_border = Border(bottom=Side(border_style="thin", color="000000"),left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"))
        no_border = Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"))

        for row in range(start_row, end_row + 1):
            # Determine the border style based on the row number
            if row % 3 == 1:  # e.g., Row 10, 13, 16, etc. (1, 4, 7, ...)
                border_style = top_border
            elif row % 3 == 2:  # e.g., Row 11, 14, 17, etc. (2, 5, 8, ...)
                border_style = no_border
            elif row % 3 == 0:  # e.g., Row 12, 15, 18, etc. (3, 6, 9, ...)
                border_style = bottom_border

            # Apply the selected border style to each cell in the specified columns
            for col in columns:
                cell = sheet[f'{col}{row}']
                cell.border = border_style


    def adjust_column_widths(self,sheet, column_widths):
        for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width
    #Formula Adder

    def addcolor(self,sheet):
            sheet['C5'].fill = PatternFill(start_color='87CEFA',end_color='87CEFA',fill_type='solid')
    def make_bold(self,sheet):
        bold_font = Font(bold=True)
        cells=['A4','A5','A6','D1','D4','D6','C5']
        for cell in cells:
            sheet[cell].font=bold_font
    def adjust_alignment(self,sheet):
        alignment = Alignment(horizontal='center', vertical='center')
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = alignment	
    def generateExcel(self,sample_data,date,v=1):
        file_name = date+'.xlsx' 
        bold_font = Font(bold=True)

        wb = Workbook()
        ws1 = wb.active
        #FIRST SHEET
        ws1.title = 'MOTHER'
        ws1['D1'].font=bold_font
        merge_ranges = ['A1:B3','C1:C3','D1:H3','A4:B4','C4','D4:E4','F4:H4','A5:B5','C5:H5','A6:B7','C6:C7','D6:E7','F6:H7','A8:B9', 'C8:C9','D8:E9','F8:H9','A10:A12','B10:B12','C10:C12','D10:E12','F10:F12','G10:G12','H10:H12']
        static_data=['Type of Vehicle:','MARUTI ECO','PPP ambulance / Empanelled vehicle (strike off which is not applicable)','Name of the Block:','GALSI – II','District:','BURDWAN:','Name of the NGO:','MOTHER','Name of the operator:','TARUN KUMAR CHATTOPADHYAY','Vehicle No.','WB-41J7245','','','YEAR:',date,'Sl. No.','Name of the Beneficiaries','Address','Type of vouchers','Distance travelled (KM.)','Date of travel','Amount claimed (Rs.)']
        i=0
        for cell in merge_ranges:
            if(":" in cell):
                ws1.merge_cells(cell)
                ws1[cell.split(":")[0]]=static_data[i]
            else:
                ws1[cell]=static_data[i]
            i+=1


        n=len(sample_data)

        start=13 #
    #    [1, 'Deepali Lohar', 'Morsidpur', 1, 22, '1-8-24', 350.00],

        distance_per_voucher=[]
        for i in range(n):
            ws1[f'A${start+1}']=i+1
            ws1[f'B${start+1}']=sample_data[i]['name'].upper()
            ws1[f'C${start+1}']=sample_data[i]['village']
            ws1[f'C${start+2}']="Burdwan"
           
            ws1[f'D${start}']="V1"
            ws1[f'D${start+1}']="V2"  
            ws1[f'D${start+2}']="V3"
            print(sample_data[i]['voucher_type'],type(sample_data[i]['voucher_type']))

            if(isinstance(sample_data[i]['voucher_type'], list)):
                for j in range(len(sample_data[i]['voucher_type'])):
                    ws1[f'F${start+j}']=sample_data[i]['distance'][j]
                    ws1[f'G${start+j}']=sample_data[i]['date'][j]
                    distance_per_voucher.append(sample_data[i]['distance'][j])

            

            else:
                distance_per_voucher.append(sample_data[i]['distance'])
                ws1[f"F${start+int(sample_data[i]['voucher_type'][1])-1}"]=sample_data[i]['distance']
                ws1[f"G${start+int(sample_data[i]['voucher_type'][1])-1}"]=sample_data[i]['date']


            start+=3
        #Add font style
        self.make_bold(ws1)
        
        self.addcolor(ws1)
        self.add_borders(ws1, start_row=10, end_row=ws1.max_row, columns=['A', 'B', 'C'])
        self.add_borders_top(ws1, 10, ws1.max_row,['D','E','F','G','H'])
        self.add_borders_top(ws1)
        #Add formula
        self.add_formula_to_column(ws1,13)
        self.add_formula_to_column2(ws1,13)
        self.add_formula_to_column3(ws1,13)
        self. adjust_column_widths(ws1,{'A':8,'B':30,'C':28,'D':8,'E':15,'F':20,'G':15,'H':20})

        # Set the alignment for each cell in the sheet
        self.adjust_alignment(ws1)
        #adjust_row_heights(ws1)

        ws2 = wb.create_sheet(title='Sheet2')
        ws2['B4'] = 'Voucher SL No'
        ws2['C4'] = 'KM PER VOUCHER'
        ws2['D4'] = 'BILL AMT PER VOUCHER Rs.'
        s=5
        for i in range(len(distance_per_voucher)):
            ws2[f'B${s+i}']=v+i
            ws2[f'C${s+i}']=distance_per_voucher[i]
        self.adjust_column_widths(ws2,{'B':15,'C':17,'D':26})
        self.add_formula_to_column2(ws2, 5,'C',4)
        self.adjust_alignment(ws2)
        self.add_borders_top(ws2,4,ws2.max_row,['B','C','D'])
        self.add_formula_to_column3(ws2,5,'D')
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)  # Rewind the buffer to the beginning

        return excel_file  # This can be sent to the frontend

        wb.save(file_name)
    def add_formula_to_column(self,sheet, start_row=13):
        """Add formula to column E based on comparison of columns F and H starting from a specific row."""
        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=5)  # Column E is the 5th column
            cell.value = f'=IF(F{row}<>H{row},"√"," ")'

    def add_formula_to_column2(self,sheet, start_row=13,col='F',col_place=8):
        """Add formula to column E based on comparison of columns F and H starting from a specific row."""
        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_place) 
            cell.value = f'=IF({col}{row}>55,{col}{row}*8,IF({col}{row}>30,450,IF({col}{row}>20,350,IF({col}{row}>10,250,IF({col}{row}>0,150,IF({col}{row}<=0,0))))))'

    def add_formula_to_column3(self,sheet, start_row=13,col_place='H'):
        """Add formula to column E based on comparison of columns F and H starting from a specific row."""
        max_row = sheet.max_row
        formula = f'=SUM({col_place}{start_row}:{col_place}{max_row})'
        
        # Add the formula to the cell in column H of the next row after the last data row
        sheet[f'{col_place}{max_row + 1}'] = formula

